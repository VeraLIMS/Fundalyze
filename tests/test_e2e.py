"""End-to-end integration tests for report generation."""
import pytest
pytest.skip("Deprecated after Directus migration", allow_module_level=True)
import pandas as pd
from openpyxl import load_workbook

import modules.generate_report.report_generator as rg
import modules.generate_report.excel_dashboard as ed
import modules.management.portfolio_manager.portfolio_manager as pm


class Dummy:
    def __init__(self, df):
        self._df = df

    def to_df(self):
        return self._df


def test_report_generation_end_to_end(tmp_path, monkeypatch):
    profile_df = pd.DataFrame({
        "symbol": ["AAA"],
        "sector": ["Tech"],
        "industry": ["Software"],
    })
    price_df = pd.DataFrame({
        "Date": pd.date_range("2023-01-01", periods=2),
        "Open": [1, 2],
        "High": [1, 2],
        "Low": [1, 2],
        "Close": [1, 2],
        "Adj Close": [1, 2],
        "Volume": [10, 20],
    })
    stmt_df = pd.DataFrame(
        {"Revenue": [100], "EPS": [5]}, index=pd.Index(["2023"], name="Period")
    )

    class FakeEquity:
        def __init__(self):
            class _Profile:
                def __call__(self, symbol):
                    return Dummy(profile_df)

            class _Price:
                def historical(self, symbol, period, provider=None):
                    return Dummy(price_df)

            class _Fundamental:
                def income(self, symbol, period):
                    return Dummy(stmt_df)

                def balance(self, symbol, period):
                    return Dummy(stmt_df)

                def cash(self, symbol, period):
                    return Dummy(stmt_df)

            self.profile = _Profile()
            self.price = _Price()
            self.fundamental = _Fundamental()

    class FakeOBB:
        def __init__(self):
            self.equity = FakeEquity()

    monkeypatch.setattr(rg, "obb", FakeOBB())

    rg.fetch_and_compile("AAA", base_output=str(tmp_path))
    ticker_dir = tmp_path / "AAA"
    assert (ticker_dir / "profile.csv").is_file()
    assert (ticker_dir / "1mo_prices.csv").is_file()
    assert (ticker_dir / "income_annual.csv").is_file()

    dash = ed.create_dashboard(output_root=str(tmp_path))
    assert dash.is_file()

    wb = load_workbook(dash)
    try:
        assert "Profile" in wb.sheetnames
        assert "PriceHistory" in wb.sheetnames
    finally:
        wb.close()


def test_portfolio_manager_cli_end_to_end(tmp_path, monkeypatch):
    data = {
        "Ticker": "AAA",
        "Name": "Alpha",
        "Sector": "Tech",
        "Industry": "Software",
        "Current Price": 10.0,
        "Market Cap": 100,
        "PE Ratio": 20.0,
        "Dividend Yield": 0.01,
    }
    monkeypatch.setattr(pm, "PORTFOLIO_FILE", str(tmp_path / "portfolio.xlsx"))
    monkeypatch.setattr(pm, "fetch_from_yfinance", lambda t: data)

    inputs = iter([
        "2",       # choose Add ticker
        "AAA",     # ticker list
        "y",       # accept info
        "5",       # exit
    ])
    monkeypatch.setattr("builtins.input", lambda *_args: next(inputs))

    pm.main()

    df = pd.read_excel(pm.PORTFOLIO_FILE)
    assert list(df["Ticker"]) == ["AAA"]
